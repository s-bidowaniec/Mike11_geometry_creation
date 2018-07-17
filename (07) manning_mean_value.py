import glob
from classes import *
from functions import *
import xlsxwriter

path = r"E:\!!Modele_IsokII\Zlotna_161152\Geodezja\txt"

XS_txt = glob.glob(path+"\*.txt")
XS_base = []
for XS_raw in XS_txt:
    with open(XS_raw, 'r') as f:
        XS_base.append(XS_t(f))
    f.close()

from openpyxl import Workbook

book = Workbook()
sheet = book.active
poz = 1
nwkData = read_NWK(open(r'C:\!!Modele ISOKII\!Etap1\161152_ZLOTNA\03_S01_Zlotna\01_MIKE11\S01_Zlotna\02_S01_Zlotna_NWK\S01_Zlotna_Q0.2%.nwk11'))
pkt = nwkData.pointList
for xs in XS_base:
    xs.get_avarage_manning()
    xs.get_km(pkt)
    print(xs.km)
    sheet.cell(row=poz, column=1, value=xs.name)
    sheet.cell(row=poz, column=2, value=xs.avManning)
    sheet.cell(row=poz, column=3, value=xs.km)
    poz += 1

book.save('avManning.xlsx')
book.close()