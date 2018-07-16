import glob
from classes import *
#from functions import *
import xlsxwriter

path = r"E:\!!Modele_IsokII\Zlotna_161152\Geodezja\txt"

XS_txt = glob.glob(path+"\*.txt")
XS_base = []
for XS_raw in XS_txt:
    with open(XS_raw, 'r') as f:
        XS_base.append(XS_t(f))
    f.close()

from openpyxl import Workbook
import time

book = Workbook()
sheet = book.active
poz = 1

for xs in XS_base:
    xs.get_avarage_manning()
    sheet.cell(row=poz, column=1, value=xs.name)
    sheet.cell(row=poz, column=2, value=xs.avManning)
    poz += 1

book.save('avManning.xlsx')
book.close()