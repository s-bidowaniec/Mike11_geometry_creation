from dbfread import DBF
import openpyxl
# wejscie, pliki z wynikami i przekroje dbf:
db = DBF(r'C:\!!Mode ISOKII\!Etap1\Oddanie.V02\1_Warstwy\BUDKOWICZANKA_S01_XS.dbf')
wb = openpyxl.load_workbook(r'C:\!!Mode ISOKII\!Etap1\Oddanie.V02\2_Wyniki\H500v3.xlsx')
# wyjsciowy xlsx do podmiany w dbf
filepath=r'C:\!!Mode ISOKII\!Etap1\Oddanie.V02\2_Wyniki\H500v3_res.xlsx'

"""
Kolumny porównywane to PROFILEM i RIVERCODE (w dbf)
arkusz wynikowy w xlsx nazwa : result (w xlsx)
kolejnosc kolumn:

nazwa rzeki, km, 010, 100, 500
"""

ws = wb['result']
db2 = []
for row in db:
    for row1 in ws.rows:
        #print(row['ProfileM'],"---",row1[1].value)
        try:
            print(int(row1[1].value))
        except:
            continue
        if abs(int(row['PROFILEM']) - int(row1[1].value)) < 2 and str(row['RIVERNAME']).lower() == str(row1[0].value).lower():
            row['010'] = row1[2].value
            row['100'] = row1[3].value
            row['500'] = row1[4].value
            print('*')

    db2.append(row)

wb = openpyxl.Workbook()

ws = wb.active

i = y = 1
for row in db2:
    for element in row.keys():
        ws.cell(row=i, column=y).value = element
        y += 1
    break

i = 2
y = 1
for row in db2:
    for element in row.values():
        ws.cell(row=i, column=y).value = element
        y += 1
    i += 1
    y = 1
wb.save(filepath)